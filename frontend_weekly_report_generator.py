#!/usr/bin/env python3
"""
前端周报生成脚本

用途：
  把开发人员复制来的原始周报 txt，按产品关键字整理成「前端工作周报」txt/xlsx。

常用命令：
  python3 frontend_weekly_report_generator.py --input raw.txt --start 0601 --end 0605

规则摘要：
  - 按关键字归类到 QY、QM、TH/L8、LW、RB88、非业务支持。
  - IF体育只写入 txt，不写入 xlsx。
  - xlsx 栏位固定：产品 / 项目 / 上周总结 / 本周计划 / 备注。
  - 没有下周计划时，计划栏写「1.日常维护。」。
  - 输出到指定 outputs 目录，默认当前目录。
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PRODUCTS = [
    {
        "key": "QY",
        "project": "QY（球友会）",
        "group": "综合/体育",
        "aliases": ["QY", "球友会", "本菲卡", "勒沃库森"],
        "xlsx": True,
    },
    {
        "key": "QM",
        "project": "QM（球盟会）",
        "group": "综合/体育",
        "aliases": ["QM", "球盟会"],
        "xlsx": True,
    },
    {
        "key": "TH",
        "project": "TH/L8（头号玩家）",
        "group": "综合/体育",
        "aliases": ["TH", "L8", "头号玩家"],
        "xlsx": True,
    },
    {
        "key": "LW",
        "project": "LW（乐玩）",
        "group": "综合/体育",
        "aliases": ["LW", "乐玩"],
        "xlsx": True,
    },
    {
        "key": "RB88",
        "project": "RB88（走地皇）",
        "group": "综合/体育",
        "aliases": ["RB88", "走地皇"],
        "xlsx": True,
    },
    {
        "key": "IF体育",
        "project": "IF体育",
        "group": "综合/体育",
        "aliases": ["IF体育", "IF 体育", "internal", "注单模块", "赛程模块", "公告模块", "结算模块"],
        "xlsx": False,
    },
    {
        "key": "飞鱼94Chat",
        "project": "飞鱼 / 94Chat",
        "group": "非业务支持",
        "aliases": ["飞鱼", "94Chat", "site-dolphin"],
        "xlsx": True,
    },
    {
        "key": "咪乐",
        "project": "咪乐",
        "group": "非业务支持",
        "aliases": ["咪乐"],
        "xlsx": True,
    },
    {
        "key": "优客服",
        "project": "优客服",
        "group": "非业务支持",
        "aliases": ["优客服", "ukefu"],
        "xlsx": True,
    },
    {
        "key": "Ezpay",
        "project": "Ezpay",
        "group": "非业务支持",
        "aliases": ["EZPAY", "Ezpay", "ezpay", "电子回单"],
        "xlsx": True,
    },
]


SUPPORT_ORDER = ["优客服", "Ezpay", "飞鱼94Chat", "咪乐"]


PLAN_MARKERS = [
    "下周计划",
    "下周进度",
    "本周计划",
    "下周工作",
    "下周计画",
]


DROP_LINES = {
    "上周总结",
    "本周工作总结",
    "周总结",
    "例行任务",
    "计划任务",
    "任务类型",
    "周报事项",
    "备注",
    "职位",
    "前端组",
    "日期",
    "截止日期",
}


@dataclass
class ProductReport:
    key: str
    project: str
    group: str
    xlsx: bool
    summary: list[str] = field(default_factory=list)
    plan: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)


def normalize_line(line: str) -> str:
    line = line.strip()
    line = line.replace("\u200e", "")
    line = re.sub(r"^[\s\-*•]+", "", line)
    line = re.sub(r"^\d+[.、]\s*", "", line)
    line = re.sub(r"^(NEW|fix|Fix|FIX)\s*[:：]\s*", "", line)
    line = re.sub(r"\s+", " ", line)
    return line.strip()


def is_dev_line(line: str) -> bool:
    return bool(re.match(r"^Dev\s+[A-Za-z0-9_-]+$", line.strip(), re.I))


def is_plan_marker(line: str) -> bool:
    return any(marker in line for marker in PLAN_MARKERS)


def split_blocks(text: str) -> list[str]:
    chunks: list[list[str]] = [[]]
    for raw in text.splitlines():
        line = raw.rstrip()
        chunks[-1].append(line)
        if is_dev_line(line):
            chunks.append([])
    return ["\n".join(chunk).strip() for chunk in chunks if "\n".join(chunk).strip()]


def product_by_key(key: str) -> dict:
    return next(item for item in PRODUCTS if item["key"] == key)


def detect_product(block: str) -> str:
    hits: list[tuple[int, str]] = []
    for item in PRODUCTS:
        score = sum(block.lower().count(alias.lower()) for alias in item["aliases"])
        if score:
            hits.append((score, item["key"]))
    if hits:
        hits.sort(reverse=True)
        return hits[0][1]

    lower = block.lower()
    if "世界杯" in block and any(word in lower for word in ["db", "pp", "openapp", "skeleton", "choice"]):
        return "LW"
    if "pa改名" in lower or "小白投注" in block:
        return "QM"
    return "QY"


def parse_block(block: str) -> tuple[str, list[str], list[str]]:
    product = detect_product(block)
    summary: list[str] = []
    plan: list[str] = []
    in_plan = False

    for raw in block.splitlines():
        line = normalize_line(raw)
        if not line:
            continue
        if is_dev_line(line):
            continue
        if is_plan_marker(line):
            in_plan = True
            continue
        if line in DROP_LINES:
            continue
        if re.fullmatch(r"\d{4}[-./]\d{1,2}[-./]\d{1,2}", line):
            continue
        if re.fullmatch(r"\d{4}[-./]\d{1,2}[-./]\d{1,2}\s*[-~至]\s*\d{4}[-./]\d{1,2}[-./]\d{1,2}", line):
            continue
        if line in ["一", "二", "三", "PC：", "H5：", "已上线：", "测试中未上线：", "持续优化项目:"]:
            continue

        target = plan if in_plan else summary
        if line not in target:
            target.append(line)

    return product, summary, plan


def numbered(items: Iterable[str]) -> str:
    clean = [item.rstrip("。") + "。" for item in items if item.strip()]
    if not clean:
        clean = ["日常维护。"]
    return "\n".join(f"{idx}.{item}" for idx, item in enumerate(clean, start=1))


def build_reports(text: str) -> dict[str, ProductReport]:
    reports: dict[str, ProductReport] = {}
    for item in PRODUCTS:
        reports[item["key"]] = ProductReport(
            key=item["key"],
            project=item["project"],
            group=item["group"],
            xlsx=item["xlsx"],
        )

    for block in split_blocks(text):
        key, summary, plan = parse_block(block)
        reports[key].summary.extend(item for item in summary if item not in reports[key].summary)
        reports[key].plan.extend(item for item in plan if item not in reports[key].plan)

    return reports


def create_txt(reports: dict[str, ProductReport], start: str, end: str, output_dir: Path) -> Path:
    title = f"前端工作周报（{start[:2]}/{start[2:]}-{end[:2]}/{end[2:]}）"
    lines = [title, "", "一、本周工作汇总"]

    active = [report for report in reports.values() if report.summary]
    for idx, report in enumerate(active, start=1):
        lines.append(f"{idx}.{report.project}：{report.summary[0].rstrip('。')}。")

    lines.append("")
    lines.append("二、按产品整理")
    for key in ["QY", "QM", "TH", "LW", "RB88", "IF体育"]:
        report = reports[key]
        if not report.summary:
            continue
        lines.append("")
        lines.append(f"【{report.project}】")
        lines.append(numbered(report.summary))
        lines.append("下周计划：")
        lines.append(numbered(report.plan))

    lines.append("")
    lines.append("三、非业务支持事项")
    for key in SUPPORT_ORDER:
        report = reports[key]
        if not report.summary:
            continue
        lines.append("")
        lines.append(f"【{report.project}】")
        lines.append(numbered(report.summary))
        lines.append("下周计划：")
        lines.append(numbered(report.plan))

    lines.append("")
    lines.append("四、每日提交与实际修改文件")
    lines.append("本次整合依据手动周报内容归纳，实际提交与修改文件以各项目 Git 记录为准。")

    path = output_dir / f"{start}_{end}_前端工作周报.txt"
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def create_xlsx(reports: dict[str, ProductReport], start: str, end: str, output_dir: Path) -> tuple[Path, Path]:
    wb = Workbook()
    ws = wb.active
    ws.title = "WEB组周报汇总"

    ws.merge_cells("A1:E1")
    ws["A1"] = "WEB组周报汇总"
    ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["产品", "项目", "上周总结", "本周计划", "备注"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row = 3
    business_start = row
    for key in ["QY", "QM", "TH", "LW", "RB88"]:
        report = reports[key]
        ws.cell(row=row, column=1, value="综合/体育" if row == business_start else "")
        ws.cell(row=row, column=2, value=report.project)
        ws.cell(row=row, column=3, value=numbered(report.summary))
        ws.cell(row=row, column=4, value=numbered(report.plan))
        ws.cell(row=row, column=5, value="\n".join(report.notes))
        row += 1

    support_start = row
    for key in SUPPORT_ORDER:
        report = reports[key]
        ws.cell(row=row, column=1, value="非业务支持" if row == support_start else "")
        ws.cell(row=row, column=2, value=report.project)
        ws.cell(row=row, column=3, value=numbered(report.summary))
        ws.cell(row=row, column=4, value=numbered(report.plan))
        ws.cell(row=row, column=5, value="\n".join(report.notes))
        row += 1

    ws.merge_cells(start_row=business_start, start_column=1, end_row=support_start - 1, end_column=1)
    ws.merge_cells(start_row=support_start, start_column=1, end_row=row - 1, end_column=1)

    for r in range(2, row):
        max_lines = 1
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.font = cell.font.copy(name="Microsoft YaHei", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.value:
                max_lines = max(max_lines, str(cell.value).count("\n") + 1)
        ws.row_dimensions[r].height = 24 if r == 2 else min(240, max(70, 18 * max_lines))

    for cell in [ws.cell(business_start, 1), ws.cell(support_start, 1)]:
        cell.font = Font(name="Microsoft YaHei", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7" if cell.row == business_start else "E2F0D9")

    for col, width in {"A": 16, "B": 20, "C": 68, "D": 42, "E": 26}.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False

    standard = output_dir / f"{start}_{end}_前端工作周报.xlsx"
    merged = output_dir / f"周报_{start}_{end}_整合版.xlsx"
    wb.save(standard)
    wb.save(merged)
    return standard, merged


def main() -> None:
    parser = argparse.ArgumentParser(description="生成前端工作周报 txt/xlsx")
    parser.add_argument("--input", "-i", required=True, help="原始周报 txt 路径")
    parser.add_argument("--start", required=True, help="开始日期，格式 MMDD，例如 0601")
    parser.add_argument("--end", required=True, help="结束日期，格式 MMDD，例如 0605")
    parser.add_argument("--output-dir", "-o", default="outputs", help="输出目录，默认 outputs")
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    text = input_path.read_text(encoding="utf-8")
    reports = build_reports(text)

    txt = create_txt(reports, args.start, args.end, output_dir)
    standard, merged = create_xlsx(reports, args.start, args.end, output_dir)

    print(txt)
    print(standard)
    print(merged)


if __name__ == "__main__":
    main()
