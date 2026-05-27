#!/usr/bin/env python3
"""Merge Outlook weekly-report next plans into the frontend weekly xlsx.

Usage:
  python3 merge_outlook_next_plan.py \
    --xlsx /path/to/0518_0522_前端工作周报.xlsx \
    --outlook-text /path/to/outlook_weekly_reports.txt \
    --output /path/to/0518_0522_前端工作周报_合并版.xlsx

This script keeps the original workbook format:
  产品 / 项目 / 上周总结 / 本周计划

The "本周计划" column is filled from Outlook "下周计划" sections by product.
"""

from __future__ import annotations

import argparse
import copy
import re
import subprocess
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


PEOPLE = {
    "AaronY": ["AaronY", "Dev AaronY", "DEV AaronY"],
    "Ayea": ["Ayea", "Dev Ayea", "DEV Ayea"],
    "Jairo": ["Jairo", "Dev Jairo", "DEV Jairo"],
    "Avram": ["Avram", "Dev Avram", "DEV Avram"],
    "Mason": ["Mason", "Dev Mason", "DEV Mason"],
    "Papa": ["Papa", "Dev Papa", "DEV Papa"],
    "Belly": ["Belly", "Dev Belly", "DEV Belly"],
    "Nolan": ["Nolan", "Dev Nolan", "DEV Nolan"],
}

COMMIT_PRODUCT_REPOS = {
    "飞鱼": Path("/Users/oncechen/IdeaProjects/feiyu-site-clean"),
    "94Chat": Path("/Users/oncechen/IdeaProjects/site-dolphin"),
}

SUMMARY_OVERRIDES = {
    "优客服": [
        "完成客服端登录接口对接，登录成功后写入 token、roles、orgi、userid 等用户信息。",
        "优化前端接口环境配置，统一 axios baseURL，并保留 mock 接口本地调试能力。",
        "新增 AGENTS.md 项目代理配置文件，规范 AI agent 协作规则和项目范围。",
        "推进体育投注用户端 ty_client 重构，拆分 SportsView 中的详情模式、协议转换、轮询和赔率闪烁逻辑。",
        "清理旧路由、无用状态和重复接口 helper，降低后续维护成本。",
        "修正投注相关 API 路径与服务前缀，并将 mock、商户 ID、网关地址改为环境变量配置。",
    ],
}


DIRECT_PRODUCT = {
    "AaronY": "QY（球友会）",
    "Avram": "QM（球盟会）",
    "Mason": "LW（乐玩）",
    "Papa": "L8（头号玩家）",
    "Belly": "RB88（走地皇）",
    "Nolan": "优客服",
}


PLAN_HEADINGS = [
    "下周工作计划",
    "下周计划",
    "下周进度",
    "本周计划",
    "计划任务",
    "Next Week",
    "next week",
]


STOP_MARKERS = [
    "Best regards",
    "Regards",
    "Dev ",
    "DEV ",
    "汇报人",
    "姓名",
]


def normalize_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("\u2028", "\n").replace("\u2029", "\n")
    text = text.replace("\ufeff", "")
    return text


def person_from_line(line: str) -> str | None:
    stripped = line.strip().strip("：:")
    for person, aliases in PEOPLE.items():
        if stripped.lower() in {alias.lower() for alias in aliases}:
            return person
        if any(re.fullmatch(rf"(Dev|DEV)\s+{re.escape(alias)}", stripped, flags=re.I) for alias in aliases):
            return person
    return None


def split_blocks(text: str) -> dict[str, str]:
    """Split pasted Outlook text into person blocks.

    Supports both common paste styles:
    - report body followed by "Dev Ayea"
    - "Dev Ayea" followed by report body
    """
    lines = normalize_text(text).splitlines()
    markers: list[tuple[int, str]] = []
    for idx, line in enumerate(lines):
        person = person_from_line(line)
        if person:
            markers.append((idx, person))

    blocks: dict[str, list[str]] = defaultdict(list)
    if not markers:
        return {}

    marker_indexes = {idx for idx, _ in markers}
    for marker_pos, (idx, person) in enumerate(markers):
        prev_end = markers[marker_pos - 1][0] + 1 if marker_pos > 0 else 0
        next_start = markers[marker_pos + 1][0] if marker_pos + 1 < len(markers) else len(lines)

        before = [line for i, line in enumerate(lines[prev_end:idx], prev_end) if i not in marker_indexes]
        after = [line for i, line in enumerate(lines[idx + 1:next_start], idx + 1) if i not in marker_indexes]

        # Outlook email bodies are usually pasted with the sender signature at
        # the end. If there is content before the person marker, prefer it.
        before_text = "\n".join(before).strip()
        after_text = "\n".join(after).strip()
        if before_text:
            blocks[person].append(before_text)
        elif after_text:
            blocks[person].append(after_text)

    return {person: "\n\n".join(parts) for person, parts in blocks.items()}


def clean_plan_line(line: str) -> str:
    line = line.strip()
    line = re.sub(r"^[\-\*•]+\s*", "", line)
    line = re.sub(r"^\d+[、.)．]\s*", "", line)
    line = re.sub(r"^[一二三四五六七八九十]+[、.)．]\s*", "", line)
    line = re.sub(r"^\d+\s+", "", line)
    return line.strip()


def extract_next_plan(block: str) -> list[str]:
    lines = [line.strip() for line in normalize_text(block).splitlines()]
    candidates: list[tuple[int, str]] = []
    for idx, line in enumerate(lines):
        if any(heading.lower() in line.lower() for heading in PLAN_HEADINGS):
            matched = next(heading for heading in PLAN_HEADINGS if heading.lower() in line.lower())
            candidates.append((idx, matched))
    if not candidates:
        return []

    preferred = [item for item in candidates if item[1] != "计划任务"]
    if preferred:
        start = preferred[0][0] + 1
    else:
        # Outlook table weekly reports may use "计划任务" as the only next-plan
        # marker. Use the last occurrence to avoid completed-task tables.
        start = candidates[-1][0] + 1

    items: list[str] = []
    for line in lines[start:]:
        stripped = line.strip()
        if not stripped:
            continue
        if any(stripped.startswith(marker) for marker in STOP_MARKERS):
            break
        if person_from_line(stripped):
            break
        if (
            "任务类型" in stripped
            or "周报事项" in stripped
            or stripped in {"备注", "计划任务", "例行任务"}
            or re.fullmatch(r"[一二三四五六七八九十]+[、.)．]\s*【.+】", stripped)
        ):
            continue
        cleaned = clean_plan_line(stripped)
        if cleaned and cleaned not in items:
            items.append(cleaned)
    return items


def route_plan(person: str, plans: list[str]) -> dict[str, list[str]]:
    routed: dict[str, list[str]] = defaultdict(list)
    if not plans:
        return routed

    if person in DIRECT_PRODUCT:
        routed[DIRECT_PRODUCT[person]].extend(plans)
        return routed

    if person == "Ayea":
        for item in plans:
            if re.search(r"EZPAY|Ezpay|内嵌", item, re.I):
                routed["Ezpay"].append(item)
        return routed

    if person == "Jairo":
        for item in plans:
            if "咪乐" in item:
                routed["咪乐"].append(item)
            else:
                routed["QY（球友会）"].append(item)
        return routed

    return routed


def dedupe(items: list[str]) -> list[str]:
    result: list[str] = []
    for item in items:
        cleaned = item.strip()
        if cleaned and cleaned not in result:
            result.append(cleaned)
    return result


def numbered(items: list[str]) -> str:
    values = dedupe(items)
    if not values:
        return "1. 日常维护"
    return "\n".join(f"{idx}. {item}" for idx, item in enumerate(values, 1))


def apply_summary_overrides(ws) -> None:
    for row in range(3, ws.max_row + 1):
        project = ws.cell(row, 2).value
        if not project:
            continue
        items = SUMMARY_OVERRIDES.get(str(project).strip())
        if not items:
            continue
        ws.cell(row, 3).value = numbered(items)
        ws.cell(row, 3).alignment = Alignment(vertical="top", wrap_text=True)


def previous_work_week(today: date | None = None) -> tuple[date, date]:
    value = today or date.today()
    monday = value - timedelta(days=value.weekday() + 7)
    friday = monday + timedelta(days=4)
    return monday, friday


def collect_commit_subjects(repo: Path, start: date, end: date) -> list[str]:
    if not (repo / ".git").exists():
        return []
    result = subprocess.run(
        [
            "git",
            "log",
            f"--since={start.isoformat()} 00:00:00",
            f"--until={end.isoformat()} 23:59:59",
            "--pretty=format:%s",
        ],
        cwd=repo,
        check=False,
        text=True,
        capture_output=True,
    )
    if result.returncode != 0:
        return []
    subjects = []
    for line in result.stdout.splitlines():
        cleaned = line.strip()
        if cleaned and not cleaned.lower().startswith("merge"):
            subjects.append(cleaned)
    return dedupe(subjects)


def plan_from_commit_product(product: str, subjects: list[str]) -> list[str]:
    if not subjects:
        return ["日常维护"]

    text = "\n".join(subjects).lower()
    items: list[str] = []
    if product == "飞鱼":
        if "apk" in text or "download" in text or "下载" in text:
            append_item(items, "飞鱼下载页、APK 与 fallback 链路日常维护")
        if "crawler" in text or "爬虫" in text:
            append_item(items, "前端爬虫遮蔽与下载安全加固")
    if product == "94Chat":
        if "apk" in text or "download" in text or "下载" in text:
            append_item(items, "94Chat 下载页与 APK 版本维护")
        if "crawler" in text or "爬虫" in text:
            append_item(items, "94Chat 前端爬虫遮蔽与下载安全加固")
    return items or ["日常维护"]


def summary_from_commit_product(product: str, subjects: list[str]) -> list[str]:
    if not subjects:
        return ["日常维护"]
    text = "\n".join(subjects).lower()
    items: list[str] = []

    if product == "飞鱼":
        if "download" in text or "下载" in text or "baidu" in text:
            append_item(items, "下载页安全加固，补充百度兜底下载与无 JS / 无 CSS 场景下的原生下载按钮。")
        if "crawler" in text or "爬虫" in text:
            append_item(items, "优化前端爬虫遮蔽策略，减少异常访问与被采集风险。")
        if "apk" in text:
            append_item(items, "同步 APK 版本资源，并维护 fallback 下载链路。")
    elif product == "94Chat":
        if "download" in text or "下载" in text or "baidu" in text:
            append_item(items, "官网页下载安全加固，补充百度兜底下载与原生下载按钮，降低异常环境影响下载入口的风险。")
        if "feiyu download hardening" in text:
            append_item(items, "同步飞鱼下载加固逻辑到 94Chat 站点，优化外部下载链路处理。")
        if "crawler" in text or "爬虫" in text:
            append_item(items, "优化前端爬虫遮蔽策略，减少异常访问与被采集风险。")
        if "apk" in text:
            append_item(items, "更新 94Chat APK 版本资源。")
        if "copyright" in text:
            append_item(items, "调整 footer copyright 年份。")
        if "ignore" in text or "ide" in text or "macos" in text:
            append_item(items, "补充 IDE 与 macOS 本地文件忽略规则，减少无关文件进入版本库。")

    return items or ["日常维护"]


def append_item(items: list[str], value: str) -> None:
    if value and value not in items:
        items.append(value)


def find_project_row(ws, project_name: str) -> int | None:
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, 2).value
        if not value:
            continue
        if str(value).strip() == project_name:
            return row
    return None


def last_content_row(ws) -> int:
    for row in range(ws.max_row, 0, -1):
        if any(ws.cell(row, col).value is not None for col in range(1, 5)):
            return row
    return ws.max_row


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy.copy(source._style)
        target.number_format = source.number_format
        target.protection = copy.copy(source.protection)
        target.alignment = copy.copy(source.alignment)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def append_product_row(ws, project_name: str, plans: list[str], category: str = "综合/体育") -> None:
    row_idx = last_content_row(ws) + 1
    ws.cell(row_idx, 1).value = category
    ws.cell(row_idx, 2).value = project_name
    ws.cell(row_idx, 3).value = "1. Outlook 周报提取补充"
    ws.cell(row_idx, 4).value = numbered(plans)
    for cell in ws[row_idx]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[row_idx].height = 120


def normalize_special_rows(ws) -> None:
    combined = find_project_row(ws, "飞鱼 / 94Chat")
    if combined is None:
        return
    start, end = previous_work_week()
    feiyu_summary = summary_from_commit_product("飞鱼", collect_commit_subjects(COMMIT_PRODUCT_REPOS["飞鱼"], start, end))
    chat_summary = summary_from_commit_product("94Chat", collect_commit_subjects(COMMIT_PRODUCT_REPOS["94Chat"], start, end))
    ws.cell(combined, 2).value = "飞鱼"
    ws.cell(combined, 3).value = numbered(feiyu_summary)
    if find_project_row(ws, "94Chat") is None:
        ws.insert_rows(combined + 1)
        copy_row_style(ws, combined, combined + 1)
        ws.cell(combined + 1, 1).value = None
        ws.cell(combined + 1, 2).value = "94Chat"
        ws.cell(combined + 1, 3).value = numbered(chat_summary)
        ws.cell(combined + 1, 4).value = "1.日常维护"


def merge_plans(xlsx: Path, outlook_text: Path, output: Path) -> None:
    blocks = split_blocks(outlook_text.read_text(encoding="utf-8"))
    product_plans: dict[str, list[str]] = defaultdict(list)

    for person, block in blocks.items():
        plans = extract_next_plan(block)
        for product, items in route_plan(person, plans).items():
            product_plans[product].extend(items)

    start, end = previous_work_week()
    for product, repo in COMMIT_PRODUCT_REPOS.items():
        subjects = collect_commit_subjects(repo, start, end)
        product_plans[product].extend(plan_from_commit_product(product, subjects))

    wb = load_workbook(xlsx)
    ws = wb.active
    normalize_special_rows(ws)
    apply_summary_overrides(ws)

    for product, plans in product_plans.items():
        row = find_project_row(ws, product)
        if row is None:
            append_product_row(ws, product, plans, category="非业务支持" if product in {"飞鱼", "94Chat"} else "综合/体育")
            continue
        ws.cell(row, 4).value = numbered(plans)
        ws.cell(row, 4).alignment = Alignment(vertical="top", wrap_text=True)

    for row in range(3, ws.max_row + 1):
        if not ws.cell(row, 2).value:
            continue
        value = ws.cell(row, 4).value
        if value is None or not str(value).strip():
            ws.cell(row, 4).value = "1.日常维护"
            ws.cell(row, 4).alignment = Alignment(vertical="top", wrap_text=True)

    last_row = last_content_row(ws)
    if ws.max_row > last_row:
        ws.delete_rows(last_row + 1, ws.max_row - last_row)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Merge Outlook next plans into frontend weekly xlsx.")
    parser.add_argument("--xlsx", required=True, type=Path, help="Source weekly xlsx.")
    parser.add_argument("--outlook-text", required=True, type=Path, help="Pasted Outlook weekly reports txt.")
    parser.add_argument("--output", required=True, type=Path, help="Output merged xlsx.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    merge_plans(args.xlsx.expanduser(), args.outlook_text.expanduser(), args.output.expanduser())
    print(args.output)


if __name__ == "__main__":
    main()
